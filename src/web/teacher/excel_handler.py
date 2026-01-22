# -*- coding: utf-8 -*-
"""
Excel导入导出处理模块
用于批量导入学生评价数据和导出报表
"""
from typing import List, Dict, Any, Optional
import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelHandler:
    """Excel文件处理器"""
    
    def __init__(self):
        if OPENPYXL_AVAILABLE:
            # 表头样式
            self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            self.header_font = Font(color="FFFFFF", bold=True, size=11)
            self.header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 边框样式
            self.thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def create_evaluation_template(
        self,
        students: List[Dict],
        indicator_name: str,
        indicator_type: str,
        options: Optional[List[str]] = None,
        max_score: float = 100
    ) -> bytes:
        """
        创建评价数据导入模板
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl库未安装，请运行: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "评价数据导入"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        
        # 表头
        headers = ['学号', '姓名', '性别', f'{indicator_name}', '备注']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # 填充学生信息
        for row_idx, student in enumerate(students, 2):
            ws.cell(row=row_idx, column=1, value=student['student_no']).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=student['name']).border = self.thin_border
            ws.cell(row=row_idx, column=3, value='男' if student['gender'] == 'male' else '女').border = self.thin_border
            
            # 评价值单元格（留空供填写）
            value_cell = ws.cell(row=row_idx, column=4, value='')
            value_cell.border = self.thin_border
            value_cell.alignment = Alignment(horizontal="center")
            
            # 备注单元格
            ws.cell(row=row_idx, column=5, value='').border = self.thin_border
        
        # 添加说明
        note_row = len(students) + 3
        ws.cell(row=note_row, column=1, value="填写说明：").font = Font(bold=True, color="FF0000")
        ws.cell(row=note_row+1, column=1, value=f"1. 请在「{indicator_name}」列填写评价数据")
        
        if indicator_type == 'score':
            ws.cell(row=note_row+2, column=1, value=f"2. 分数范围：0-{max_score}")
        elif indicator_type == 'level':
            ws.cell(row=note_row+2, column=1, value=f"2. 可选等级：{', '.join(options) if options else '优秀/良好/及格/不及格'}")
        elif indicator_type == 'boolean':
            ws.cell(row=note_row+2, column=1, value="2. 请填写：是 或 否")
        
        ws.cell(row=note_row+3, column=1, value="3. 备注栏为可选项")
        ws.cell(row=note_row+4, column=1, value="4. 请勿修改学号、姓名等基础信息")
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def parse_evaluation_import(self, file_content: bytes) -> List[Dict[str, Any]]:
        """
        解析导入的评价数据Excel
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl库未安装")
        
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                continue
            student_no, name, gender, value = row[:4]
            remark = row[4] if len(row) > 4 else ''
            
            if not student_no or value is None or value == '':
                continue
            
            data.append({
                'student_no': str(student_no).strip(),
                'name': str(name).strip() if name else '',
                'value': str(value).strip(),
                'remark': str(remark).strip() if remark else ''
            })
        
        return data
    
    def export_class_evaluations(
        self,
        class_name: str,
        semester_name: str,
        students_data: List[Dict[str, Any]]
    ) -> bytes:
        """
        导出班级评价数据报表
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl库未安装")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "班级评价报表"
        
        # 标题
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"{class_name} {semester_name} 综合素质评价报表"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # 生成时间
        ws.merge_cells('A2:F2')
        time_cell = ws['A2']
        time_cell.value = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        time_cell.font = Font(size=10, color="666666")
        time_cell.alignment = Alignment(horizontal="center")
        
        # 表头
        if students_data:
            headers = ['学号', '姓名', '性别']
            sample_evaluations = students_data[0].get('evaluations', [])
            for eval_item in sample_evaluations:
                headers.append(eval_item['indicator_name'])
            headers.append('综合评分')
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.header_alignment
                cell.border = self.thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(header)) + 2)
            
            # 填充数据
            for row_idx, student_data in enumerate(students_data, 5):
                ws.cell(row=row_idx, column=1, value=student_data['student_no']).border = self.thin_border
                ws.cell(row=row_idx, column=2, value=student_data['name']).border = self.thin_border
                ws.cell(row=row_idx, column=3, value='男' if student_data.get('gender') == 'male' else '女').border = self.thin_border
                
                col_idx = 4
                total_score = 0
                count = 0
                for eval_item in student_data.get('evaluations', []):
                    value = eval_item.get('value', '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal="center")
                    
                    try:
                        score = float(value)
                        total_score += score
                        count += 1
                    except (ValueError, TypeError):
                        level_scores = {'优秀': 95, '良好': 85, '及格': 70, '不及格': 50}
                        if value in level_scores:
                            total_score += level_scores[value]
                            count += 1
                    
                    col_idx += 1
                
                avg_score = round(total_score / count, 1) if count > 0 else 0
                score_cell = ws.cell(row=row_idx, column=col_idx, value=avg_score)
                score_cell.border = self.thin_border
                score_cell.font = Font(bold=True, color="0000FF")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_student_comments(
        self,
        class_name: str,
        semester_name: str,
        comments_data: List[Dict[str, Any]]
    ) -> bytes:
        """
        导出班级评语
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl库未安装")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "期末评语"
        
        # 标题
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"{class_name} {semester_name} 期末评语"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # 表头
        headers = ['学号', '姓名', 'AI评语', '教师评语']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
        
        # 填充数据
        for row_idx, comment in enumerate(comments_data, 4):
            ws.cell(row=row_idx, column=1, value=comment['student_no']).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=comment['student_name']).border = self.thin_border
            
            ai_cell = ws.cell(row=row_idx, column=3, value=comment.get('ai_comment', ''))
            ai_cell.border = self.thin_border
            ai_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            teacher_cell = ws.cell(row=row_idx, column=4, value=comment.get('teacher_comment', ''))
            teacher_cell.border = self.thin_border
            teacher_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            ws.row_dimensions[row_idx].height = 60
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


# 全局实例
excel_handler = ExcelHandler()
